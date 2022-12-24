import numpy
import os
import openpyxl
import pandas as pd
import numpy as np
import torch
import torch.nn as nn
from torch.utils.data import Dataset,DataLoader
class Model(nn.Module):
    def __init__(self, input_dim):
        super(Model, self).__init__()
        self.model = nn.Sequential(
            nn.Linear(input_dim, 64),
            nn.ReLU(),
            nn.Linear(64, 8),
            nn.Linear(8, 1)
        )
        self.loss = nn.MSELoss()

    def forward(self, x):
        x = self.model(x)
        x=x.squeeze(1)
        return x

    def cal_loss(self,outputs, targets):
        train_loss = self.loss.forward(outputs, targets)
        return train_loss
class Covid19DataSet(Dataset):

    def __init__(self, x, y=None):
        if y is None:
            self.y = y
        else:
            self.y = torch.Tensor(np.array(y))
        self.x = torch.Tensor(np.array(x))

    def __getitem__(self, item):
        if self.y is None:
            return self.x[item]
        else:
            return self.x[item], self.y[item]

    def __len__(self):
        return len(self.x)

test_df=pd.read_csv('D:/PythonProjects/MachineLeaning/HW1/Data/covid_test.csv')
z=test_df[test_df.columns[1:94]]
test_dataset=Covid19DataSet(z)
test_DataLoader=DataLoader(test_dataset,batch_size=256,drop_last=False)
def save_test_pred(input_dim,test_dataloader):
    model=Model(input_dim)
    model.load_state_dict(torch.load("covid_model.pth"))
    pred_list=[]
    id_list=[]
    id=0
    for data in test_dataloader:
        preds=model(data)
        for pred in preds:
            pred_list.append(float(pred))
            id_list.append(int(id))
            id+=1
    os.chdir("..")
    work_book=openpyxl.Workbook()
    work_sheet=work_book.create_sheet('Submission')
    row=1
    column=1
    for id,tested_positive in enumerate(pred_list):

        work_sheet.cell(row,column).value=id
        column+=1
        work_sheet.cell(row,column).value=tested_positive
        row+=1
        column=1
    del work_book["Sheet"]
    work_book.save('Submission.csv')

save_test_pred(z.shape[1],test_DataLoader)
